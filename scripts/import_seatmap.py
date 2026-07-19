"""Import the real hall layout from the color-coded Excel into the database.

Source: ``[NHH 2026] Sơ đồ phòng hòa nhạc lớn.xlsx`` (repo root), tab ``Sơ đồ``.

Every blue/peach/green cell is one seat. Tier comes from the fill color; the seat
number is the cell's value; the (section, row_label) is derived from the cell's
position in the grid; svg_x / svg_y come from the spreadsheet column / row so the
rendered map matches the Excel.

Run:  .venv/bin/python -m scripts.import_seatmap
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string as col_idx
from sqlalchemy import delete

from app.db import SessionLocal
from app.models import OrderItem, PriceTier, Seat, Ticket

REPO_ROOT = Path(__file__).resolve().parent.parent
SHEET = "Sơ đồ"


def _find_xlsx() -> Path:
    """Locate the seat-map workbook in the repo root by glob.

    The filename contains Vietnamese characters; matching it as a literal string
    breaks across filesystems that normalize Unicode differently (macOS vs Linux
    containers). Globbing sidesteps that, and is independent of the working dir.
    """
    matches = sorted(REPO_ROOT.glob("*.xlsx"))
    if not matches:
        raise FileNotFoundError(f"No .xlsx seat map found in {REPO_ROOT}")
    return matches[0]

# Excel fill color (ARGB) -> tier definition (display name, price in VND). The Excel
# source colors (blue/peach/green) only *detect* which tier a cell belongs to. The
# UI colour is NOT stored — it's derived from price rank in the front-end palette.
TIERS = {
    "FFD0F0FF": ("Sông Trời", 700_000),
    "FFF7D6C8": ("Dòng Chảy", 500_000),
    "FFE2F5ED": ("Mạch Nguồn", 300_000),
}

# Center aisle / row-label column, and key wing columns (by spreadsheet letter).
AF = col_idx("AF")          # 32 — center column holding the row letters
COL_C = col_idx("C")        # Tầng 3 side, left
COL_F, COL_G = col_idx("F"), col_idx("G")        # Tầng 2 left: L2B, L2A
COL_BG, COL_BH = col_idx("BG"), col_idx("BH")    # Tầng 2 right: R2A, R2B
COL_BK = col_idx("BK")      # Tầng 3 side, right

# Legend swatches live in the top-left block; never treat those as real seats.
LEGEND_MAX_COL = col_idx("M")
LEGEND_MAX_ROW = 12


def _seat_color(cell) -> str | None:
    f = cell.fill
    if not (f and f.fgColor and f.fgColor.type == "rgb"):
        return None
    rgb = f.fgColor.rgb
    return rgb if rgb in TIERS else None


def classify(col: int, row: int, row_letters: dict[int, str]) -> tuple[str, str]:
    """Return (section, row_label) for a seat at the given grid position."""
    if col in (COL_F, COL_G):
        return "Tầng 2", ("L2A" if col == COL_G else "L2B")
    if col in (COL_BG, COL_BH):
        return "Tầng 2", ("R2A" if col == COL_BG else "R2B")
    if col == COL_C:
        return "Tầng 3", "L"
    if col == COL_BK:
        return "Tầng 3", "R"
    # Center block: top rows (<=12) are Tầng 3, the main floor is Tầng 1.
    section = "Tầng 3" if row <= 12 else "Tầng 1"
    return section, row_letters.get(row, "?")


def run() -> None:
    wb = openpyxl.load_workbook(_find_xlsx(), data_only=True)
    ws = wb[SHEET]

    # Row letter for each spreadsheet row, read from the center label column (AF).
    row_letters: dict[int, str] = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, AF).value
        if isinstance(v, str) and v.strip():
            row_letters[r] = v.strip()

    db = SessionLocal()
    try:
        # Clean slate (children first for FKs).
        db.execute(delete(Ticket))
        db.execute(delete(OrderItem))
        db.execute(delete(Seat))
        db.execute(delete(PriceTier))
        db.flush()

        tiers: dict[str, PriceTier] = {}
        for rgb, (name, price) in TIERS.items():
            t = PriceTier(name=name, price_vnd=price)
            db.add(t)
            tiers[rgb] = t
        db.flush()

        seen: set[tuple[str, str, int]] = set()
        counts = {name: 0 for name, _ in TIERS.values()}
        skipped: list[str] = []

        for row in ws.iter_rows():
            for cell in row:
                rgb = _seat_color(cell)
                if not rgb:
                    continue
                c, r = cell.column, cell.row
                # Exclude the legend swatches in the top-left corner.
                if r <= LEGEND_MAX_ROW and c <= LEGEND_MAX_COL and c < col_idx("N"):
                    continue
                num = cell.value
                if not isinstance(num, (int, float)):
                    skipped.append(f"{cell.coordinate} (no number)")
                    continue
                num = int(num)
                section, row_label = classify(c, r, row_letters)
                key = (section, row_label, num)
                if key in seen:
                    skipped.append(f"{cell.coordinate} dup {key}")
                    continue
                seen.add(key)

                label = f"{section} – Hàng {row_label} – Ghế {num}"
                db.add(
                    Seat(
                        section=section,
                        row_label=row_label,
                        seat_number=num,
                        label=label,
                        tier_id=tiers[rgb].id,
                        svg_x=float(c),
                        svg_y=float(r),
                        status="available",
                    )
                )
                counts[tiers[rgb].name] += 1

        db.commit()
        total = sum(counts.values())
        print(f"Imported {total} seats:")
        for name, n in counts.items():
            print(f"  {name}: {n}")
        if skipped:
            print(f"Skipped {len(skipped)} cell(s):")
            for s in skipped:
                print(f"  - {s}")
    finally:
        db.close()


if __name__ == "__main__":
    run()
