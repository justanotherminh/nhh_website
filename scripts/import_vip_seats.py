"""Seed the VIP / invitation pool from scripts/data/vip_reserved_seats.csv.

This is a **first-boot seed only**. The pool's source of truth is ``seats.is_vip``
in the database, edited by managers at ``/admin/vip-seats``. Re-applying the CSV on
every deploy would silently undo their changes, so ``run()`` does nothing once any
seat is already marked VIP. Use ``--force`` to override that deliberately.

A seeded seat gets ``is_vip=True`` and ``status='blocked'``: removed from public sale
(the seat map and the hold API both require 'available'), but still issuable as a free
invitation via the admin "Vé mời" page.

The reserved list was extracted from the greyed-out cells of the masterplan's
"Sơ đồ hạng vé" tab (145 seats). Regenerate it with --regen <file.xlsx> if that
map changes (dev only; the workbook is not committed).

Run inside the app container:
    python -m scripts.import_vip_seats            # seed, if the pool is empty
    python -m scripts.import_vip_seats --force    # seed even if the pool is set up
    python -m scripts.import_vip_seats --unblock  # release the whole pool to sale
"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

from sqlalchemy import func, select, update

from app.db import SessionLocal
from app.models import Seat

CSV_PATH = Path(__file__).resolve().parent / "data" / "vip_reserved_seats.csv"


def _load_list() -> list[tuple[str, str, int]]:
    with open(CSV_PATH, newline="", encoding="utf-8") as fh:
        return [
            (r["section"], r["row_label"], int(r["seat_number"]))
            for r in csv.DictReader(fh)
        ]


def _regen(xlsx: str) -> None:
    """Dev-only: rebuild the CSV from a masterplan workbook's greyed cells."""
    import openpyxl

    from scripts.import_seatmap import AF, classify

    ws = openpyxl.load_workbook(xlsx, data_only=True)["Sơ đồ hạng vé"]
    row_letters = {
        r: ws.cell(r, AF).value.strip()
        for r in range(1, ws.max_row + 1)
        if isinstance(ws.cell(r, AF).value, str) and ws.cell(r, AF).value.strip()
    }
    seats = set()
    for row in ws.iter_rows():
        for cell in row:
            f = cell.fill
            rgb = f.fgColor.rgb if (f and f.fgColor and f.fgColor.type == "rgb") else None
            if rgb == "FFCCCCCC" and isinstance(cell.value, (int, float)):
                sec, rl = classify(cell.column, cell.row, row_letters)
                seats.add((sec, rl, int(cell.value)))
    rows = sorted(seats)
    with open(CSV_PATH, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["section", "row_label", "seat_number"])
        w.writerows(rows)
    print(f"Regenerated {CSV_PATH} with {len(rows)} reserved seats.")


def run(unblock: bool = False, force: bool = False) -> None:
    db = SessionLocal()
    try:
        if unblock:
            # Escape hatch: hand the whole pool back to public sale. Exported
            # invitations are 'booked' and are left alone — releasing those would
            # put a seat with a live QR back on sale.
            released = db.execute(
                update(Seat)
                .where(Seat.is_vip.is_(True), Seat.status == "blocked")
                .values(status="available", is_vip=False,
                        held_by_cart=None, hold_expires_at=None)
                .returning(Seat.id)
            ).scalars().all()
            kept = db.execute(
                select(func.count()).select_from(Seat).where(Seat.is_vip.is_(True))
            ).scalar()
            db.commit()
            print(f"Released {len(released)} VIP seats back to sale.")
            if kept:
                print(f"  {kept} left VIP (already exported/booked, untouched).")
            return

        existing = db.execute(
            select(func.count()).select_from(Seat).where(Seat.is_vip.is_(True))
        ).scalar()
        if existing and not force:
            # The pool is managed in the admin UI now; re-seeding would revert it.
            print(f"VIP pool already initialised ({existing} seats) -> skipping seed.")
            return

        reserved = _load_list()
        changed = booked = missing = 0
        for sec, rl, num in reserved:
            seat = db.execute(
                select(Seat).where(
                    Seat.section == sec, Seat.row_label == rl, Seat.seat_number == num
                )
            ).scalar_one_or_none()
            if seat is None:
                missing += 1
                print(f"  ! not found: {sec} – {rl} – {num}")
                continue
            if seat.status == "booked":
                booked += 1  # never touch a sold seat
                print(f"  ! already booked, skipped: {seat.label}")
                continue
            if seat.status == "available":
                db.execute(
                    update(Seat).where(Seat.id == seat.id).values(
                        status="blocked", is_vip=True,
                        held_by_cart=None, hold_expires_at=None,
                    )
                )
                changed += 1
        db.commit()
        print(f"Seeded {changed}/{len(reserved)} VIP seats (available -> blocked).")
        if booked:
            print(f"  {booked} already booked (left untouched).")
        if missing:
            print(f"  {missing} not found in the seat map.")
    finally:
        db.close()


if __name__ == "__main__":
    args = sys.argv[1:]
    if args and args[0] == "--regen":
        _regen(args[1])
    else:
        run(unblock="--unblock" in args, force="--force" in args)
